from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from app.modules.dashboard import obtener_resumen_dashboard


def exportar_dashboard_excel():
    resumen = obtener_resumen_dashboard()

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Dashboard"

    # Título
    hoja.merge_cells("A1:B1")
    hoja["A1"] = "GM7 Dashboard Profesional"
    hoja["A1"].font = Font(bold=True, size=16)
    hoja["A1"].alignment = Alignment(horizontal="center")

    # Encabezados
    encabezados = ["Indicador", "Valor"]

    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=3, column=columna)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

    datos = [
        ["Clientes activos", resumen["clientes_activos"]],
        ["Cotizaciones activas", resumen["cotizaciones_activas"]],
        ["Órdenes activas", resumen["ordenes_activas"]],
        ["Diseños activos", resumen["disenos_activos"]],
        ["Producción activa", resumen["produccion_activa"]],
        ["Pagos registrados", resumen["pagos_registrados"]],
        ["Saldo pendiente", resumen["saldo_pendiente"]],
    ]

    for fila, registro in enumerate(datos, start=4):
        for columna, valor in enumerate(registro, start=1):
            celda = hoja.cell(row=fila, column=columna)
            celda.value = valor

            celda.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

    hoja["B10"].number_format = '"$"#,##0.00'

    hoja.column_dimensions["A"].width = 25
    hoja.column_dimensions["B"].width = 18

    libro.save("GM7_Dashboard_Real.xlsx")

    print("Dashboard exportado correctamente.")


if __name__ == "__main__":
    exportar_dashboard_excel()