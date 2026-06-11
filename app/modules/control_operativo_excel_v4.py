from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.chart.marker import DataPoint

PASSWORD = "gm7"

COLOR_AZUL_OSCURO = "1A3A5C"
COLOR_AZUL_MEDIO = "2E5F8A"
COLOR_AZUL_CLARO = "EBF1F8"
COLOR_VERDE_OSCURO = "1E6B3C"
COLOR_VERDE_CLARO = "E8F5EE"
COLOR_GRIS = "BFBFBF"
COLOR_BLANCO = "FFFFFF"

FILA_INICIO = 3
FILA_FIN = 202
TOTAL_FILAS = 200


def borde(color=COLOR_GRIS, estilo="thin"):
    lado = Side(style=estilo, color=color)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def encabezado(celda, color=COLOR_AZUL_OSCURO):
    celda.font = Font(name="Arial", bold=True, color=COLOR_BLANCO, size=10)
    celda.fill = PatternFill("solid", start_color=color)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = borde(COLOR_AZUL_OSCURO, "medium")


def proteger(ws):
    ws.protection.sheet = True
    ws.protection.password = PASSWORD


def desbloquear(ws, rango):
    for fila in ws[rango]:
        for celda in fila:
            celda.protection = Protection(locked=False)


def construir_trabajos(wb):
    ws = wb.active
    ws.title = "TRABAJOS"

    encabezados = [
        "ID Trabajo", "Fecha Pedido", "Cliente", "Tipo Trabajo", "Descripción",
        "Cantidad", "¿Trae Diseño?", "Costo Diseño", "Total Trabajo",
        "Anticipo 50%", "Anticipo Recibido", "Saldo Pendiente",
        "Responsable Diseño", "Fecha Compromiso", "Fecha Entrega Real",
        "Prioridad", "Status", "Pago Completo", "Comentarios"
    ]

    anchos = [12, 14, 24, 18, 35, 10, 14, 14, 14, 14, 14, 14, 20, 15, 15, 14, 18, 14, 35]

    ws.merge_cells("A1:S1")
    ws["A1"] = "GM7 PRO · CONTROL OPERATIVO DE TRABAJOS"
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", start_color=COLOR_AZUL_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, titulo in enumerate(encabezados, start=1):
        c = ws.cell(row=2, column=col, value=titulo)
        encabezado(c, COLOR_AZUL_MEDIO)
        ws.column_dimensions[get_column_letter(col)].width = anchos[col - 1]

    for fila in range(FILA_INICIO, FILA_FIN + 1):
        fill = COLOR_AZUL_CLARO if fila % 2 == 0 else COLOR_BLANCO

        for col in range(1, 20):
            c = ws.cell(row=fila, column=col)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=fill)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = borde()

        ws.cell(row=fila, column=1).value = f'=IF(B{fila}="","","FC-"&TEXT(ROW()-2,"0000"))'
        ws.cell(row=fila, column=10).value = f'=IF(I{fila}="","",I{fila}*0.5)'
        ws.cell(row=fila, column=12).value = f'=IF(I{fila}="","",I{fila}-K{fila})'
        ws.cell(row=fila, column=18).value = f'=IF(I{fila}="","",IF(L{fila}<=0,"Sí","No"))'

        for col in [8, 9, 10, 11, 12]:
            ws.cell(row=fila, column=col).number_format = '"$"#,##0.00'

    columnas_editables = ["B", "C", "D", "E", "F", "G", "H", "I", "K", "M", "N", "O", "P", "Q", "S"]
    for col in columnas_editables:
        desbloquear(ws, f"{col}{FILA_INICIO}:{col}{FILA_FIN}")

    validaciones = [
        ('"Sí,No"', f"G{FILA_INICIO}:G{FILA_FIN}"),
        ('"Baja,Media,Alta,Prioritaria"', f"P{FILA_INICIO}:P{FILA_FIN}"),
        ('"Pendiente,En diseño,En producción,Listo para entrega,Entregado,Cancelado"', f"Q{FILA_INICIO}:Q{FILA_FIN}"),
    ]

    for lista, rango in validaciones:
        dv = DataValidation(type="list", formula1=lista, allow_blank=True)
        dv.sqref = rango
        ws.add_data_validation(dv)
        # Formato condicional para Pago Completo
    # ==============================
    # Formato condicional Pago Completo
    # ==============================

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    dxf_verde = DifferentialStyle(fill=verde)
    dxf_rojo = DifferentialStyle(fill=rojo)

    regla_verde = Rule(type="containsText", operator="containsText",
                    text="Liquidado", dxf=dxf_verde)

    regla_roja = Rule(type="containsText", operator="containsText",
                    text="Por cobrar", dxf=dxf_rojo)

    ws.conditional_formatting.add(
        f"R{FILA_INICIO}:R{FILA_FIN}",
        regla_verde
    )

    ws.conditional_formatting.add(
        f"R{FILA_INICIO}:R{FILA_FIN}",
        regla_roja
    )
    fila_total = FILA_FIN + 1
    ws.merge_cells(f"A{fila_total}:H{fila_total}")
    ws[f"A{fila_total}"] = "TOTALES →"
    ws[f"A{fila_total}"].font = Font(bold=True, color=COLOR_AZUL_OSCURO)
    ws[f"A{fila_total}"].alignment = Alignment(horizontal="right")

    for col in ["I", "J", "K", "L"]:
        c = ws[f"{col}{fila_total}"]
        c.value = f"=SUM({col}{FILA_INICIO}:{col}{FILA_FIN})"
        c.font = Font(bold=True, color=COLOR_AZUL_OSCURO)
        c.fill = PatternFill("solid", start_color="D6E4F0")
        c.border = borde(COLOR_AZUL_OSCURO, "medium")
        c.number_format = '"$"#,##0.00'

    ws.auto_filter.ref = "A2:S202"
    ws.freeze_panes = "B3"
    proteger(ws)


def construir_pagos(wb):
    ws = wb.create_sheet("PAGOS")

    headers = [
        "ID Trabajo", "Fecha Pedido", "Cliente", "Descripción",
        "Total Trabajo", "Anticipo Recibido", "Saldo Pendiente", "Estado Pago"
    ]

    ws.merge_cells("A1:H1")
    ws["A1"] = "GM7 PRO · ESTADO DE CUENTA Y PAGOS"
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", start_color=COLOR_VERDE_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, titulo in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=titulo)
        encabezado(c, COLOR_VERDE_OSCURO)
        ws.column_dimensions[get_column_letter(col)].width = 18

    for i, fila_t in enumerate(range(FILA_INICIO, FILA_FIN + 1), start=FILA_INICIO):
        fill = COLOR_VERDE_CLARO if i % 2 == 0 else COLOR_BLANCO

        formulas = [
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!A{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!B{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!C{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!E{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!I{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!K{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",TRABAJOS!L{fila_t})',
            f'=IF(TRABAJOS!B{fila_t}="","",IF(TRABAJOS!L{fila_t}<=0,"Liquidado","Parcial"))',
        ]

        for col, formula in enumerate(formulas, start=1):
            c = ws.cell(row=i, column=col, value=formula)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=fill)
            c.border = borde()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in [5, 6, 7]:
            ws.cell(row=i, column=col).number_format = '"$"#,##0.00'

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = "A2:H202"
    proteger(ws)


def kpi(ws, fila, col, titulo, formula, color, formato=None):
    c1 = ws.cell(row=fila, column=col, value=titulo)
    c1.font = Font(bold=True, color=COLOR_BLANCO)
    c1.fill = PatternFill("solid", start_color=color)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c1.border = borde(color, "medium")

    c2 = ws.cell(row=fila, column=col + 1, value=formula)
    c2.font = Font(bold=True, size=14, color=color)
    c2.fill = PatternFill("solid", start_color="F5F9FF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = borde(color, "medium")
    if formato:
        c2.number_format = formato


def construir_dashboard(wb):
    ws = wb.create_sheet("DASHBOARD")

    ws.merge_cells("A1:N1")
    ws["A1"] = "GM7 PRO · DASHBOARD EJECUTIVO"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", start_color=COLOR_AZUL_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center")

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.merge_cells("A3:J3")
    ws["A3"] = "RESUMEN FINANCIERO"
    encabezado(ws["A3"], COLOR_AZUL_OSCURO)

    kpi(ws, 4, 1, "Total Ventas", f'=SUM(TRABAJOS!I{FILA_INICIO}:I{FILA_FIN})', "1A3A5C", '"$"#,##0.00')
    kpi(ws, 4, 3, "Anticipos", f'=SUM(TRABAJOS!K{FILA_INICIO}:K{FILA_FIN})', "1E6B3C", '"$"#,##0.00')
    kpi(ws, 4, 5, "Saldo Pendiente", f'=SUM(TRABAJOS!L{FILA_INICIO}:L{FILA_FIN})', "C55A11", '"$"#,##0.00')
    kpi(ws, 4, 7, "Liquidados", f'=COUNTIF(TRABAJOS!R{FILA_INICIO}:R{FILA_FIN},"Sí")', "2E5F8A")
    kpi(ws, 4, 9, "Por Cobrar", f'=COUNTIF(TRABAJOS!R{FILA_INICIO}:R{FILA_FIN},"No")', "C00000")

    ws.merge_cells("A7:J7")
    ws["A7"] = "ESTADO DE PRODUCCIÓN"
    encabezado(ws["A7"], COLOR_AZUL_MEDIO)

    kpi(ws, 8, 1, "Pendientes", f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"Pendiente")', "757171")
    kpi(ws, 8, 3, "En Diseño", f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"En diseño")', "4472C4")
    kpi(ws, 8, 5, "En Producción", f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"En producción")', "ED7D31")
    kpi(ws, 8, 7, "Listos", f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"Listo para entrega")', "FFC000")
    kpi(ws, 8, 9, "Entregados", f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"Entregado")', "70AD47")

    ws.merge_cells("A11:J11")
    ws["A11"] = "DISEÑO Y PRIORIDADES"
    encabezado(ws["A11"], "7030A0")

    kpi(ws, 12, 1, "Trae Diseño", f'=COUNTIF(TRABAJOS!G{FILA_INICIO}:G{FILA_FIN},"Sí")', "7030A0")
    kpi(ws, 12, 3, "Requiere Diseño", f'=COUNTIF(TRABAJOS!G{FILA_INICIO}:G{FILA_FIN},"No")', "B4009E")
    kpi(ws, 12, 5, "Prioritarios", f'=COUNTIF(TRABAJOS!P{FILA_INICIO}:P{FILA_FIN},"Prioritaria")', "C55A11")
    kpi(ws, 12, 7, "Alta Prioridad", f'=COUNTIF(TRABAJOS!P{FILA_INICIO}:P{FILA_FIN},"Alta")', "ED7D31")
    kpi(ws, 12, 9, "Total Trabajos", f'=COUNTIF(TRABAJOS!B{FILA_INICIO}:B{FILA_FIN},"<>")', "1A3A5C")

    crear_tablas_graficas(ws)
    crear_graficas(ws)

    proteger(ws)


def crear_tablas_graficas(ws):
    ws["P1"] = "Status"
    ws["Q1"] = "Cantidad"

    status = [
        "Pendiente", "En diseño", "En producción",
        "Listo para entrega", "Entregado", "Cancelado"
    ]

    for i, s in enumerate(status, start=2):
        ws.cell(row=i, column=16, value=s)
        ws.cell(row=i, column=17, value=f'=COUNTIF(TRABAJOS!Q{FILA_INICIO}:Q{FILA_FIN},"{s}")')

    ws["P10"] = "Pago"
    ws["Q10"] = "Cantidad"
    ws["P11"] = "Liquidado"
    ws["Q11"] = f'=COUNTIF(TRABAJOS!R{FILA_INICIO}:R{FILA_FIN},"Sí")'
    ws["P12"] = "Por cobrar"
    ws["Q12"] = f'=COUNTIF(TRABAJOS!R{FILA_INICIO}:R{FILA_FIN},"No")'

    ws["P15"] = "Finanzas"
    ws["Q15"] = "Monto"
    ws["P16"] = "Ventas"
    ws["Q16"] = f'=SUM(TRABAJOS!I{FILA_INICIO}:I{FILA_FIN})'
    ws["P17"] = "Anticipos"
    ws["Q17"] = f'=SUM(TRABAJOS!K{FILA_INICIO}:K{FILA_FIN})'
    ws["P18"] = "Saldo"
    ws["Q18"] = f'=SUM(TRABAJOS!L{FILA_INICIO}:L{FILA_FIN})'

    ws.column_dimensions["P"].hidden = False
    ws.column_dimensions["Q"].hidden = False


def crear_graficas(ws):
    chart1 = BarChart()
    chart1.title = "Trabajos por Status"
    chart1.height = 10
    chart1.width = 11
    chart1.add_data(Reference(ws, min_col=17, min_row=2, max_row=7), titles_from_data=False)
    chart1.set_categories(Reference(ws, min_col=16, min_row=2, max_row=7))
    ws.add_chart(chart1, "A16")

    chart2 = PieChart()
    chart2.title = "Pago Completo"
    chart2.height = 10
    chart2.width = 11
    chart2.add_data(Reference(ws, min_col=17, min_row=11, max_row=12), titles_from_data=False)
    chart2.set_categories(Reference(ws, min_col=16, min_row=11, max_row=12))
    ws.add_chart(chart2, "E16")

    chart3 = BarChart()
    chart3.title = "Resumen Financiero"
    chart3.height = 10
    chart3.width = 11
    chart3.add_data(Reference(ws, min_col=17, min_row=16, max_row=18), titles_from_data=False)
    chart3.set_categories(Reference(ws, min_col=16, min_row=16, max_row=18))
    ws.add_chart(chart3, "I16")


def construir_datos_empresa(wb):
    ws = wb.create_sheet("DATOS_EMPRESA")

    ws.merge_cells("A1:B1")
    ws["A1"] = "GM7 PRO · DATOS DE LA EMPRESA"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", start_color=COLOR_AZUL_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Campo"
    ws["B2"] = "Valor"
    encabezado(ws["A2"], COLOR_AZUL_MEDIO)
    encabezado(ws["B2"], COLOR_AZUL_MEDIO)

    datos = [
        ("Nombre empresa", "GM7 Impresión Profesional"),
        ("Razón social", "GM7 Servicios Gráficos SA de CV"),
        ("RFC", "GM701010ABC"),
        ("Teléfono", "33 1234 5678"),
        ("Correo", "contacto@gm7.com"),
        ("Dirección", "Guadalajara, Jalisco"),
        ("Mensaje cotización", "Vigencia de 15 días naturales."),
        ("Mensaje recibo", "Gracias por su preferencia."),
        ("Logo", "Pendiente de integración"),
    ]

    for fila, (campo, valor) in enumerate(datos, start=3):
        ws.cell(row=fila, column=1, value=campo)
        ws.cell(row=fila, column=2, value=valor)

        for col in [1, 2]:
            c = ws.cell(row=fila, column=col)
            c.border = borde()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.cell(row=fila, column=1).font = Font(bold=True, color=COLOR_AZUL_OSCURO)
        ws.cell(row=fila, column=1).fill = PatternFill("solid", start_color=COLOR_AZUL_CLARO)
        ws.cell(row=fila, column=2).protection = Protection(locked=False)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    proteger(ws)


def crear_control_operativo():
    wb = Workbook()

    construir_trabajos(wb)
    construir_pagos(wb)
    construir_dashboard(wb)
    construir_datos_empresa(wb)

    ruta = "GM7_Control_Operativo_V4.xlsx"
    wb.save(ruta)

    print("✅ Archivo guardado:", ruta)
    print("📋 Hojas creadas:", wb.sheetnames)
    print("📊 Filas operativas:", TOTAL_FILAS)


if __name__ == "__main__":
    crear_control_operativo()