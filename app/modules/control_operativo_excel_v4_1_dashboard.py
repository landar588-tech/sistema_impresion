# -*- coding: utf-8 -*-
"""
control_operativo_excel_v4_1_dashboard.py
===============================================================================
FullColor · CONTROL OPERATIVO — V4.1 Dashboard (orientado a decisiones)
-------------------------------------------------------------------------------
Dashboard ejecutivo diseñado para responder 8 preguntas de negocio en < 15 seg.

Estructura visual (3 zonas, 3 niveles de lectura):
    ZONA A  → 4 KPIs grandes (ventas, pendiente de cobro, atrasados, listos)
    ZONA B  → 2 donuts (distribución por status + cobranza)
    ZONA C  → 2 líneas de tendencia (ventas por semana + por mes)

Tablas auxiliares en hoja oculta AUX_DASHBOARD (no visibles en DASHBOARD).

NO modifica la V4 estable. Reutiliza por import:
    - construir_trabajos       → TRABAJOS intacta
    - construir_pagos          → PAGOS intacta
    - construir_datos_empresa  → DATOS_EMPRESA intacta

Ejecución (Windows):
    py app\\modules\\control_operativo_excel_v4_1_dashboard.py

Salida:
    FullColor_Control_Operativo_V4_1_Dashboard.xlsx
===============================================================================
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

# Reutilizamos la V4 estable sin modificarla.
import control_operativo_excel_v4 as v4

# Constantes de rango operativo (heredadas de V4).
FI = v4.FILA_INICIO   # 3
FF = v4.FILA_FIN      # 202
DINERO = '"$"#,##0.00'

# =============================================================================
# PALETA DE COLORES — mínima, intencionada
# =============================================================================
AZUL_TITULO = "1A3A5C"     # Título principal y bordes
AZUL_KPI = "2E5F8A"        # KPI Ventas
NARANJA_ALERTA = "E74C3C"  # KPI Atrasados (rojo-naranja para generar urgencia)
VERDE = "27AE60"           # Entregado / Liquidado / KPI Listos
ROJO = "C0392B"            # Por cobrar
NARANJA2 = "D35400"        # KPI Pendiente de cobro
GRIS_STATUS = "95A5A6"     # Pendiente
AZUL_DISENO = "3498DB"     # En diseño
NARANJA_PROD = "E67E22"    # En producción
AMARILLO = "F39C12"        # Listo para entrega
GRIS_CANCEL = "7F8C8D"     # Cancelado
BLANCO = "FFFFFF"
FONDO = "F8F9FA"           # Fondo general del dashboard (gris casi blanco)
FONDO_KPI = "FFFFFF"       # Fondo del valor del KPI

# Mapa de colores por sector del donut Status.
STATUS_LIST = ["Pendiente", "En diseño", "En producción",
               "Listo para entrega", "Entregado", "Cancelado"]
STATUS_COLORS = [GRIS_STATUS, AZUL_DISENO, NARANJA_PROD,
                 AMARILLO, VERDE, GRIS_CANCEL]


# =============================================================================
# UTILIDADES VISUALES (mínimas, sin ruido)
# =============================================================================
def _borde(color="D5D8DC"):
    lado = Side(style="thin", color=color)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _solido(color):
    """GraphicalProperties con relleno sólido."""
    return GraphicalProperties(solidFill=color)


def _colorear_puntos(serie, colores):
    """Asigna color a cada sector/punto de una serie de gráfica."""
    for idx, color in enumerate(colores):
        dp = DataPoint(idx=idx)
        dp.graphicalProperties = _solido(color)
        serie.data_points.append(dp)


def _kpi_grande(ws, fila, col, etiqueta, formula, color, formato=None):
    """
    Tarjeta KPI minimalista de 2 filas × 3 columnas:
        Fila 1: etiqueta (fondo de color, texto blanco, centrado)
        Fila 2: valor grande (fondo blanco, texto del color, centrado)
    Prioriza legibilidad: número GRANDE, etiqueta discreta.
    """
    span = 3  # columnas que ocupa cada tarjeta
    ultima = col + span - 1

    ws.merge_cells(start_row=fila, start_column=col,
                   end_row=fila, end_column=ultima)
    ws.merge_cells(start_row=fila + 1, start_column=col,
                   end_row=fila + 1, end_column=ultima)

    # Etiqueta
    t = ws.cell(row=fila, column=col, value=etiqueta)
    t.font = Font(name="Arial", bold=True, size=9, color=BLANCO)
    t.fill = PatternFill("solid", start_color=color)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Valor
    v = ws.cell(row=fila + 1, column=col, value=formula)
    v.font = Font(name="Arial", bold=True, size=20, color=color)
    v.fill = PatternFill("solid", start_color=FONDO_KPI)
    v.alignment = Alignment(horizontal="center", vertical="center")
    if formato:
        v.number_format = formato

    # Bordes suaves en toda la tarjeta.
    for r in (fila, fila + 1):
        for c in range(col, ultima + 1):
            ws.cell(row=r, column=c).border = _borde(color)

    ws.row_dimensions[fila].height = 18
    ws.row_dimensions[fila + 1].height = 36


# =============================================================================
# DATOS DE EJEMPLO (10 trabajos para validar que los KPIs reflejan datos reales)
# =============================================================================
def _insertar_datos_ejemplo(wb):
    """
    Inserta 10 trabajos de ejemplo directamente en la hoja TRABAJOS para que
    los KPIs y gráficas muestren valores reales al abrir el archivo.

    Se escriben SOLO las columnas editables (B,C,D,E,F,G,H,I,K,N,P,Q).
    Las columnas con fórmula (A=folio, J=anticipo50%, L=saldo, R=pago) se
    calculan automáticamente por las fórmulas de la V4.

    NO modifica la estructura, fórmulas ni protección de TRABAJOS.
    """
    ws = wb["TRABAJOS"]

    hoy = date.today()

    # (fecha, cliente, tipo, descripcion, cant, trae_diseno, costo_diseno,
    #  total, anticipo_recibido, fecha_compromiso, prioridad, status)
    ejemplos = [
        (hoy - timedelta(days=35), "Comercial López", "Lona", "Lona 3x2m evento",
         2, "No", 500, 4500, 2250, hoy - timedelta(days=28), "Alta", "Entregado"),
        (hoy - timedelta(days=30), "Restaurante El Sol", "Menú", "Menús plastificados",
         100, "Sí", 0, 3200, 3200, hoy - timedelta(days=20), "Media", "Entregado"),
        (hoy - timedelta(days=25), "Gym Power", "Vinil", "Vinil decorativo 5m",
         1, "No", 800, 6000, 3000, hoy - timedelta(days=15), "Alta", "Entregado"),
        (hoy - timedelta(days=20), "Farmacia Vida", "Tarjetas", "Tarjetas de presentación",
         500, "Sí", 0, 1800, 900, hoy - timedelta(days=10), "Baja", "Listo para entrega"),
        (hoy - timedelta(days=15), "Café Aroma", "Banner", "Banner roll up",
         3, "No", 600, 3500, 1750, hoy - timedelta(days=5), "Media", "Listo para entrega"),
        (hoy - timedelta(days=12), "Escuela Patria", "Folleto", "Folletos trípticos",
         200, "Sí", 0, 2800, 1400, hoy - timedelta(days=2), "Alta", "En producción"),
        (hoy - timedelta(days=10), "Taller Mecánico JR", "Lona", "Lona fachada 4x3",
         1, "No", 700, 5500, 2750, hoy - timedelta(days=1), "Prioritaria", "En producción"),
        (hoy - timedelta(days=7), "Boutique María", "Etiqueta", "Etiquetas adhesivas",
         1000, "No", 400, 2200, 1100, hoy + timedelta(days=3), "Media", "En diseño"),
        (hoy - timedelta(days=5), "Notaría 12", "Folder", "Folders corporativos",
         300, "Sí", 0, 4000, 2000, hoy + timedelta(days=7), "Alta", "Pendiente"),
        (hoy - timedelta(days=2), "Club Deportivo", "Poster", "Posters A2 evento",
         50, "No", 350, 1500, 750, hoy + timedelta(days=10), "Media", "Pendiente"),
    ]

    for i, datos in enumerate(ejemplos):
        fila = FI + i  # filas 3, 4, 5, ... 12
        (fecha, cliente, tipo, desc, cant, diseno, costo_d,
         total, anticipo, fecha_comp, prioridad, status) = datos

        ws.cell(row=fila, column=2, value=fecha)       # B: Fecha Pedido
        ws.cell(row=fila, column=3, value=cliente)     # C: Cliente
        ws.cell(row=fila, column=4, value=tipo)        # D: Tipo Trabajo
        ws.cell(row=fila, column=5, value=desc)        # E: Descripción
        ws.cell(row=fila, column=6, value=cant)        # F: Cantidad
        ws.cell(row=fila, column=7, value=diseno)      # G: ¿Trae Diseño?
        ws.cell(row=fila, column=8, value=costo_d)     # H: Costo Diseño
        ws.cell(row=fila, column=9, value=total)       # I: Total Trabajo
        ws.cell(row=fila, column=11, value=anticipo)   # K: Anticipo Recibido
        ws.cell(row=fila, column=14, value=fecha_comp) # N: Fecha Compromiso
        ws.cell(row=fila, column=16, value=prioridad)  # P: Prioridad
        ws.cell(row=fila, column=17, value=status)     # Q: Status

        # Formato de fecha para las columnas B y N.
        ws.cell(row=fila, column=2).number_format = "dd/mm/yyyy"
        ws.cell(row=fila, column=14).number_format = "dd/mm/yyyy"


# =============================================================================
# HOJA OCULTA: AUX_DASHBOARD (tablas que alimentan gráficas)
# =============================================================================
def construir_aux_dashboard(wb):
    """
    Crea AUX_DASHBOARD (oculta) con las tablas auxiliares:
        A1:B7   → Status (conteo por estado)
        A10:B12 → Cobranza (Liquidado / Por cobrar)
        A15:D21 → Ventas por semana (6 semanas): A=label B=Ventas C=Inicio D=Fin
        A24:D30 → Ventas por mes (6 meses): A=label B=Ventas C=Inicio D=Fin
    """
    aux = wb.create_sheet("AUX_DASHBOARD")

    # --- Status ---
    aux["A1"] = "Status"
    aux["B1"] = "Cantidad"
    for i, estado in enumerate(STATUS_LIST, start=2):
        aux.cell(row=i, column=1, value=estado)
        aux.cell(row=i, column=2,
                 value=f'=COUNTIF(TRABAJOS!Q{FI}:Q{FF},"{estado}")')

    # --- Cobranza (Liquidado / Por cobrar) ---
    aux["A10"] = "Estado"
    aux["B10"] = "Cantidad"
    aux["A11"] = "Liquidado"
    aux["B11"] = f'=COUNTIF(TRABAJOS!R{FI}:R{FF},"Sí")'
    aux["A12"] = "Por cobrar"
    aux["B12"] = f'=COUNTIF(TRABAJOS!R{FI}:R{FF},"No")'

    # --- Ventas por semana (últimas 6) ---
    # A=Semana(label) B=Ventas C=Inicio D=Fin E=Pico (solo para marcador rojo)
    aux["A15"] = "Semana"
    aux["B15"] = "Ventas"
    aux["C15"] = "Inicio"
    aux["D15"] = "Fin"
    aux["E15"] = "Pico"
    s_ini, s_fin = 16, 21
    for r in range(s_ini, s_fin + 1):
        atras = s_fin - r
        aux.cell(row=r, column=3,
                 value=f"=TODAY()-WEEKDAY(TODAY(),2)+1-7*{atras}")
        aux.cell(row=r, column=4, value=f"=C{r}+6")
        aux.cell(row=r, column=1, value=f'=TEXT(C{r},"dd/mm")')
        aux.cell(row=r, column=2,
                 value=(f'=SUMIFS(TRABAJOS!$I${FI}:$I${FF},'
                        f'TRABAJOS!$B${FI}:$B${FF},">="&C{r},'
                        f'TRABAJOS!$B${FI}:$B${FF},"<="&D{r})'))
        # Pico: solo el valor de la semana con mayor venta; el resto NA().
        aux.cell(row=r, column=5,
                 value=(f'=IF(AND(B{r}=MAX($B${s_ini}:$B${s_fin}),B{r}>0),'
                        f'B{r},NA())'))
        aux.cell(row=r, column=2).number_format = DINERO
        aux.cell(row=r, column=5).number_format = DINERO

    # --- Ventas por mes (últimos 6) ---
    # A=Mes(label) B=Ventas C=Inicio D=Fin
    aux["A24"] = "Mes"
    aux["B24"] = "Ventas"
    aux["C24"] = "Inicio"
    aux["D24"] = "Fin"
    m_ini, m_fin = 25, 30
    for r in range(m_ini, m_fin + 1):
        atras = m_fin - r
        aux.cell(row=r, column=3,
                 value=f"=DATE(YEAR(TODAY()),MONTH(TODAY())-{atras},1)")
        aux.cell(row=r, column=4, value=f"=EOMONTH(C{r},0)")
        aux.cell(row=r, column=1, value=f'=TEXT(C{r},"mmm yy")')
        aux.cell(row=r, column=2,
                 value=(f'=SUMIFS(TRABAJOS!$I${FI}:$I${FF},'
                        f'TRABAJOS!$B${FI}:$B${FF},">="&C{r},'
                        f'TRABAJOS!$B${FI}:$B${FF},"<="&D{r})'))
        aux.cell(row=r, column=2).number_format = DINERO

    # Ocultar la hoja.
    aux.sheet_state = "hidden"
    v4.proteger(aux)
    return aux


# =============================================================================
# GRÁFICAS
# =============================================================================
def _donut_status(aux, ws, ancla):
    """
    Donut de distribución por status.
    Etiquetas: SOLO porcentaje. Nombres de categoría en la leyenda.
    Entregado en verde.
    """
    ch = PieChart()
    ch.style = 26
    ch.title = "Distribución por Status"
    ch.width = 9
    ch.height = 9

    datos = Reference(aux, min_col=2, min_row=1, max_row=7)
    cats = Reference(aux, min_col=1, min_row=2, max_row=7)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)

    # Etiquetas: SOLO porcentaje. Sin nombres, sin valores.
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showPercent = True
    ch.dataLabels.showVal = False
    ch.dataLabels.showCatName = False
    ch.dataLabels.showSerName = False

    # Colores por sector (Entregado = verde).
    _colorear_puntos(ch.series[0], STATUS_COLORS)

    ws.add_chart(ch, ancla)


def _donut_cobranza(aux, ws, ancla):
    """
    Donut de cobranza: Liquidado (verde) vs Por cobrar (rojo).
    Etiquetas: SOLO nombre + porcentaje. Sin valores numéricos duplicados.
    """
    ch = PieChart()
    ch.style = 26
    ch.title = "Cobranza"
    ch.width = 9
    ch.height = 9

    datos = Reference(aux, min_col=2, min_row=10, max_row=12)
    cats = Reference(aux, min_col=1, min_row=11, max_row=12)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)

    # Etiquetas: nombre + porcentaje. Sin valor numérico (evita duplicar).
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showCatName = True
    ch.dataLabels.showPercent = True
    ch.dataLabels.showVal = False
    ch.dataLabels.showSerName = False

    _colorear_puntos(ch.series[0], [VERDE, ROJO])

    ws.add_chart(ch, ancla)


def _linea_semanal(aux, ws, ancla):
    """
    Línea de ventas por semana.
    UNA sola serie visible "Ventas". Semanas como categorías del eje X.
    Sin etiquetas en los puntos.
    Segunda serie oculta "Pico" con marcador rojo grande SOLO en la semana top.
    """
    ch = LineChart()
    ch.title = "Ventas por Semana"
    ch.width = 9
    ch.height = 8
    ch.y_axis.numFmt = '"$"#,##0'
    ch.y_axis.title = "Ventas"
    ch.x_axis.title = "Semana"

    # Serie principal: Ventas (col B)
    datos = Reference(aux, min_col=2, min_row=15, max_row=21)
    cats = Reference(aux, min_col=1, min_row=16, max_row=21)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)

    # Serie pico (col E): marcador rojo grande SOLO en la semana con más ventas.
    pico = Reference(aux, min_col=5, min_row=15, max_row=21)
    ch.add_data(pico, titles_from_data=True)

    # Estilo serie principal: línea azul, marcadores pequeños, SIN etiquetas.
    s = ch.series[0]
    s.smooth = False
    s.marker = Marker(symbol="circle", size=5)
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=AZUL_KPI, w=25000)

    # Estilo serie pico: sin línea, solo marcador rojo grande, sin etiquetas.
    sp = ch.series[1]
    sp.smooth = False
    sp.marker = Marker(symbol="circle", size=14)
    sp.marker.graphicalProperties = _solido(NARANJA_ALERTA)
    sp.graphicalProperties = GraphicalProperties()
    sp.graphicalProperties.line = LineProperties(noFill=True)

    # Ocultar serie pico de la leyenda: desactivar leyenda completa ya que
    # la serie principal es la única relevante.
    ch.legend = None

    ws.add_chart(ch, ancla)


def _linea_mensual(aux, ws, ancla):
    """
    Línea de ventas por mes.
    UNA sola serie "Ventas". Meses como categorías del eje X.
    Sin etiquetas en todos los puntos para mantener limpieza.
    """
    ch = LineChart()
    ch.title = "Ventas por Mes"
    ch.width = 9
    ch.height = 8
    ch.y_axis.numFmt = '"$"#,##0'
    ch.y_axis.title = "Ventas"
    ch.x_axis.title = "Mes"
    ch.legend = None  # Sin leyenda (una sola serie)

    # Serie única: Ventas
    datos = Reference(aux, min_col=2, min_row=24, max_row=30)
    cats = Reference(aux, min_col=1, min_row=25, max_row=30)
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)

    # Estilo: línea verde, marcadores diamante, SIN etiquetas de valor.
    s = ch.series[0]
    s.smooth = False
    s.marker = Marker(symbol="diamond", size=6)
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=VERDE, w=25000)

    ws.add_chart(ch, ancla)


# =============================================================================
# HOJA DASHBOARD (3 zonas — 3 niveles de lectura)
# =============================================================================
def construir_dashboard_v41(wb):
    """
    Construye el DASHBOARD ejecutivo + la hoja oculta AUX_DASHBOARD.
    Distribución visual idéntica a la aprobada.
    """
    ws = wb.create_sheet("DASHBOARD")
    aux = construir_aux_dashboard(wb)

    # --- Configuración del lienzo ---
    ws.sheet_view.showGridLines = False
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # =========================================================================
    # TÍTULO
    # =========================================================================
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "FULLCOLOR · DASHBOARD EJECUTIVO"
    c.font = Font(name="Arial", bold=True, size=16, color=BLANCO)
    c.fill = PatternFill("solid", start_color=AZUL_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Fila 2: separador visual (vacío).
    ws.row_dimensions[2].height = 6

    # =========================================================================
    # ZONA A — 4 KPIs GRANDES (filas 3-4)
    # =========================================================================

    # KPI 1: Ventas totales
    _kpi_grande(ws, 3, 1, "VENTAS TOTALES",
                f'=SUM(TRABAJOS!I{FI}:I{FF})',
                AZUL_KPI, DINERO)

    # KPI 2: Pendiente de cobro (SUM del saldo pendiente = dinero real)
    _kpi_grande(ws, 3, 5, "PENDIENTE DE COBRO",
                f'=SUM(TRABAJOS!L{FI}:L{FF})',
                NARANJA2, DINERO)

    # KPI 3: Atrasados (excluye Entregado y Cancelado, Fecha Compromiso < HOY)
    _kpi_grande(ws, 3, 9, "ATRASADOS",
                (f'=COUNTIFS(TRABAJOS!Q{FI}:Q{FF},"<>Entregado",'
                 f'TRABAJOS!Q{FI}:Q{FF},"<>Cancelado",'
                 f'TRABAJOS!N{FI}:N{FF},"<"&TODAY(),'
                 f'TRABAJOS!N{FI}:N{FF},"<>")'),
                NARANJA_ALERTA)

    # KPI 4: Listos para entregar
    _kpi_grande(ws, 3, 12, "LISTOS P/ENTREGAR",
                f'=COUNTIF(TRABAJOS!Q{FI}:Q{FF},"Listo para entrega")',
                VERDE)

    # Separador.
    ws.row_dimensions[5].height = 10

    # =========================================================================
    # ZONA B — 2 DONUTS (filas 6-22)
    # =========================================================================

    ws.merge_cells("A6:N6")
    st = ws["A6"]
    st.value = "DISTRIBUCIÓN OPERATIVA"
    st.font = Font(name="Arial", bold=True, size=10, color=AZUL_TITULO)
    st.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[6].height = 20

    _donut_status(aux, ws, "A8")
    _donut_cobranza(aux, ws, "H8")

    # =========================================================================
    # ZONA C — 2 LÍNEAS DE TENDENCIA (filas 24-40)
    # =========================================================================

    ws.merge_cells("A24:N24")
    st2 = ws["A24"]
    st2.value = "TENDENCIA DE VENTAS"
    st2.font = Font(name="Arial", bold=True, size=10, color=AZUL_TITULO)
    st2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[24].height = 20

    _linea_semanal(aux, ws, "A26")
    _linea_mensual(aux, ws, "H26")

    # Proteger la hoja.
    v4.proteger(ws)


# =============================================================================
# ORQUESTADOR
# =============================================================================
def crear_control_operativo_v41():
    """
    Genera el libro completo:
        - TRABAJOS, PAGOS, DATOS_EMPRESA → V4 estable (sin modificar)
        - DASHBOARD → rediseño V4.1
        - AUX_DASHBOARD → hoja oculta de soporte
        - 10 datos de ejemplo para validar KPIs
    """
    wb = Workbook()

    # Forzar recálculo completo de fórmulas al abrir el archivo.
    # Sin esto, Excel puede mostrar 0 en KPIs que leen de otras hojas.
    # fullCalcOnLoad: recalcula TODAS las fórmulas al abrir.
    # calcMode="auto": mantiene recálculo automático activo.
    # calcId alto: invalida la caché de Excel y fuerza recálculo completo.
    wb.calculation = CalcProperties(
        fullCalcOnLoad=True, calcMode="auto", calcId=999999
    )

    v4.construir_trabajos(wb)        # TRABAJOS intacta
    v4.construir_pagos(wb)           # PAGOS intacta
    construir_dashboard_v41(wb)      # DASHBOARD nuevo + AUX_DASHBOARD oculta
    v4.construir_datos_empresa(wb)   # DATOS_EMPRESA intacta

    # Insertar 10 trabajos de ejemplo DESPUÉS de construir la estructura
    # para que los KPIs reflejen datos reales al abrir el archivo.
    _insertar_datos_ejemplo(wb)

    ruta = "FullColor_Control_Operativo_V4_1_Dashboard.xlsx"
    wb.save(ruta)

    print("=" * 60)
    print("  FullColor · Control Operativo V4.1 Dashboard")
    print("=" * 60)
    print(f"  Archivo guardado : {ruta}")
    print(f"  Hojas visibles   : TRABAJOS, PAGOS, DASHBOARD, DATOS_EMPRESA")
    print(f"  Hoja oculta      : AUX_DASHBOARD")
    print(f"  Filas operativas : {v4.TOTAL_FILAS}")
    print(f"  Datos de ejemplo : 10 trabajos precargados")
    print("=" * 60)


if __name__ == "__main__":
    crear_control_operativo_v41()
